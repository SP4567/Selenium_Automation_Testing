from openpyxl import Workbook
wb = Workbook()

ws = wb.active
ws['A1'] = "Name"
wb.save('writedemo.xlsx')


testData = [['Name', 'City'],['Suyash', 'Lucknow'], ['MY', 'Lucknow']]
for data in testData:
    ws.append(data)
wb.save('demo.xlsx')

for i in range(1, 6):
    for j in range(1, 5):
        ws.cell(row=i, column=j).value = i+j
wb.save('demo2.xlsx')

#reading data from openpyxl
from openpyxl import Workbook, load_workbook
wb = load_workbook(filename='demo.xlsx')
sh = wb.active
print(sh['A3'].value)
print(wb['Demo']['A2'].value)
print(sh.cell(2,3).value)
row_ct = sh.max_row
col_ct = sh.max_column
for i in range(row_ct, row_ct + col_ct):
    for j in range(col_ct, col_ct + row_ct):
        ws.cell(row = i, column=j).value = i + j #print(sh.cell(row=i, column=j).value)