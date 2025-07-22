from openpyxl import Workbook
wb = Workbook()
ws = wb.active
from faker import Faker
fake_data = Faker()
print(fake_data.name())
print(fake_data.address())
print(fake_data.email())
print(fake_data.phone_number())
print(fake_data.url())

for i in range(100):
    print(fake_data.name(), fake_data.address(), fake_data.email(), fake_data.phone_number(), fake_data.url())

for i in range(1, 11):
    for j in range(1, 4):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=1).value = fake_data.email()
        ws.cell(row=i, column=1).value = fake_data.address()
wb.save()